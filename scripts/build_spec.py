"""
Build the SDTM Source-to-Target mapping specification (sdtm_spec.xlsx).
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

spec_data = [
    {"Domain": "DM", "SDTM Variable": "STUDYID", "Label": "Study Identifier",
     "Type": "Char", "Source Dataset": "Assigned", "Source Variable": "N/A",
     "Derivation Rule": "Hardcoded: 'CDISCPILOT01'", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "DOMAIN", "Label": "Domain Abbreviation",
     "Type": "Char", "Source Dataset": "Assigned", "Source Variable": "N/A",
     "Derivation Rule": "Hardcoded: 'DM'", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "USUBJID", "Label": "Unique Subject Identifier",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "STUDYID + SUBJID",
     "Derivation Rule": "STUDYID || '-' || SUBJID", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "SUBJID", "Label": "Subject Identifier",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "SUBJID",
     "Derivation Rule": "Direct mapping", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "SITEID", "Label": "Study Site Identifier",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "SITEID",
     "Derivation Rule": "Direct mapping", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "BRTHDTC", "Label": "Date/Time of Birth",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "BRTHDT",
     "Derivation Rule": "Convert BRTHDT to ISO 8601 (YYYY-MM-DD)", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "AGE", "Label": "Age",
     "Type": "Num", "Source Dataset": "Derived", "Source Variable": "BRTHDTC, RFSTDTC",
     "Derivation Rule": "floor((RFSTDTC - BRTHDTC) / 365.25)", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "AGEU", "Label": "Age Units",
     "Type": "Char", "Source Dataset": "Assigned", "Source Variable": "N/A",
     "Derivation Rule": "Hardcoded: 'YEARS'", "CT Codelist": "AGEU"},
    {"Domain": "DM", "SDTM Variable": "AGEGR1", "Label": "Pooled Age Group 1",
     "Type": "Char", "Source Dataset": "Derived", "Source Variable": "AGE",
     "Derivation Rule": "AGE<65 → '<65'; 65≤AGE≤80 → '65-80'; AGE>80 → '>80'",
     "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "SEX", "Label": "Sex",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "SEX",
     "Derivation Rule": "Direct mapping", "CT Codelist": "SEX"},
    {"Domain": "DM", "SDTM Variable": "RACE", "Label": "Race",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "RACE",
     "Derivation Rule": "Direct mapping", "CT Codelist": "RACE"},
    {"Domain": "DM", "SDTM Variable": "ETHNIC", "Label": "Ethnicity",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "ETHNIC",
     "Derivation Rule": "Direct mapping", "CT Codelist": "ETHNIC"},
    {"Domain": "DM", "SDTM Variable": "COUNTRY", "Label": "Country",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "COUNTRY",
     "Derivation Rule": "Direct mapping", "CT Codelist": "COUNTRY"},
    {"Domain": "DM", "SDTM Variable": "RFSTDTC", "Label": "Subject Reference Start Date/Time",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "RFSTDTC",
     "Derivation Rule": "Convert to ISO 8601", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "RFENDTC", "Label": "Subject Reference End Date/Time",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "RFENDTC",
     "Derivation Rule": "Convert to ISO 8601", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "ARMCD", "Label": "Planned Arm Code",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "ARMCD",
     "Derivation Rule": "Direct mapping", "CT Codelist": ""},
    {"Domain": "DM", "SDTM Variable": "ARM", "Label": "Description of Planned Arm",
     "Type": "Char", "Source Dataset": "enrollment.csv", "Source Variable": "ARM",
     "Derivation Rule": "Direct mapping", "CT Codelist": ""},
]

def build_spec():
    out_dir = os.path.join(os.path.dirname(__file__), "..", "specs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "sdtm_spec.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "DM Mapping"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = list(spec_data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(spec_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_data["Type"] == "Num":
                cell.font = Font(color="0000FF")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["H"].width = 14

    ws.auto_filter.ref = f"A1:H{len(spec_data)+1}"
    wb.save(out_path)
    print(f"[SPEC] SDTM specification → {out_path}")

if __name__ == "__main__":
    build_spec()
